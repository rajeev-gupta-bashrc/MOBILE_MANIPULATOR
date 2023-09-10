#!/usr/bin/env python
import rospkg
import rospy, math
from std_msgs.msg import Float64, Float64MultiArray, Bool
import openpyxl
from control_msgs.msg import JointControllerState
from geometry_msgs.msg import Twist
from sensor_msgs.msg import JointState
import time

class DifferentialDriveController:
    def __init__(self, left_joint_name, right_joint_name):
        self.cmd_vel_pub = rospy.Publisher('/cmd_vel', Twist, queue_size=10)

        # Store joint names
        self.left_joint_name = left_joint_name
        self.right_joint_name = right_joint_name

        # Define control gains (adjust as needed)
        self.kp = 0.1  # Proportional gain
        self.ki = 0.01  # Integral gain
        self.kd = 0.05  # Derivative gain

        # Initialize PID variables
        self.left_integral_error = 0.0
        self.right_integral_error = 0.0
        self.prev_left_error = 0.0
        self.prev_right_error = 0.0

    def control_joints(self, left_theta_desired, right_theta_desired):
        # Define the rate at which the while loop will run (in Hz)
        rate = rospy.Rate(5)  # Adjust this rate as needed

        while not rospy.is_shutdown():
            # Get current joint states
            joint_states = rospy.wait_for_message('/joint_states', JointState)

            # Extract left and right joint angles
            left_theta_current = joint_states.position[joint_states.name.index(self.left_joint_name)]
            right_theta_current = joint_states.position[joint_states.name.index(self.right_joint_name)]

            # Calculate the error in joint angles
            left_error = left_theta_desired - left_theta_current
            right_error = right_theta_desired - right_theta_current

            # Calculate control commands using PID controller
            self.left_integral_error += left_error
            self.right_integral_error += right_error

            left_cmd_vel = self.kp * left_error + self.ki * self.left_integral_error + self.kd * (left_error - self.prev_left_error)
            right_cmd_vel = self.kp * right_error + self.ki * self.right_integral_error + self.kd * (right_error - self.prev_right_error)

            self.prev_left_error = left_error
            self.prev_right_error = right_error

            # Create a Twist message for cmd_vel
            cmd_vel_msg = Twist()
            cmd_vel_msg.linear.x = (left_cmd_vel + right_cmd_vel) / 2.0
            cmd_vel_msg.angular.z = (-left_cmd_vel + right_cmd_vel) / 2.0

            # Publish cmd_vel
            self.cmd_vel_pub.publish(cmd_vel_msg)

            # Check if both joint angles are close to the desired values
            epsilon = 0.05  # Tolerance for angle error
            if abs(left_error) < epsilon and abs(right_error) < epsilon:
                rospy.loginfo("Reached desired wheel joint angles.")
                break

            rate.sleep()
        cmd_vel_msg = Twist()
        self.cmd_vel_pub.publish(cmd_vel_msg)
    
    def move(self, thetas):
        try:
            rospy.loginfo('Got wheel data')
            self.control_joints(thetas.data[0], thetas.data[1])
            rospy.loginfo('wheel moved, go ahead')
        except (IndexError, ValueError):
            rospy.loginfo('some error in move_bot.py -> DifferentialDriveController -> move()')
            
            
# Load the Excel file
rospack = rospkg.RosPack()
# input_file_path = rospack.get_path('tb3_manipulator')  + '/data/circle.xlsx'
# input_file_path = rospack.get_path('tb3_manipulator')  + '/data/line_data.xlsx'
input_file_path = rospack.get_path('tb3_manipulator')  + '/data/Line_Tracking.xlsx'
workbook = openpyxl.load_workbook(input_file_path)
sheet = workbook.active

if __name__ == '__main__':
    try:
        rospy.init_node('excel_to_ros_node', anonymous=True)
        rospy.loginfo('Node Initiated')
        
        arm_joint_topics = [f'/manipulator/joint{i+1}_position' for i in range(5)]
        arm_joint_pub = [rospy.Publisher(topic+'/command', Float64, queue_size=10) for topic in arm_joint_topics]
        wheel_controller = DifferentialDriveController('wheel_left_joint', 'wheel_right_joint')
        ## wheel_joint_pub = rospy.Publisher('/manipulator/wheel_theta', Float64MultiArray, queue_size=10)
        
        arm_joint_state_subs = []
        rate1 = rospy.Rate(1) 
        num_columns = sum(1 for cell in sheet[1] if cell.value is not None)
        num_rows = 7
        
        # try:
        rospy.loginfo('Cols: %d', num_columns)
        rospy.loginfo('Rows: %d', num_rows)
        
        for col in range(1, num_columns + 1):
            wheel_theta = [sheet.cell(row=1, column=col).value, sheet.cell(row=2, column=col).value]
            wheel_theta = [x*math.pi/180 for x in wheel_theta]
            
            for row in range(3, num_rows + 1):  # Rows 3 to 7
                value = sheet.cell(row=row, column=col).value
                value = Float64(value*math.pi/180)
                arm_joint_pub[row-3].publish(value)
                rospy.loginfo(f'Published {col}th value {value} on /manipulator/joint{row - 2}/command')
                
            ## wheel_joint_pub.publish(Float64MultiArray(data=wheel_theta))
            wheel_controller.control_joints(wheel_theta[0], wheel_theta[1])
            rospy.loginfo(f'{col}th iteration: wheel joints published')
            
            # check if the bot has acheived the previously passed values
            rate2 = rospy.Rate(10)
            _check = True
            while(_check):
                _check = False
                arm_states = [rospy.wait_for_message(topic+'/state', JointControllerState).error for topic in arm_joint_topics ]
                for err in arm_states:
                    if err > 0.01:
                        _check = True
                        break
                rate2.sleep()
                
            # wait if first iteration
            wt=5
            if(col==1):
                rospy.loginfo('Wait for %d secs :)', wt)
                time.sleep(wt)
            # rate1.sleep()
            
                
        rospy.spin()
        # except Exception as e:
        #     rospy.loginfo(e)
    except rospy.ROSInterruptException:
        pass


#x->y = z
#y->z = x
#z->x = y